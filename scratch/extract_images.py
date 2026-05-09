import openpyxl
import io
import os
from PIL import Image

def get_column_letter(col_idx):
    col_idx += 1
    letters = ''
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters

wb = openpyxl.load_workbook('2026 Kerala Assembly Election - Dashboard.xlsx')
ws = wb['Candidates']

os.makedirs('assets/images/candidates', exist_ok=True)

count = 0
for image in ws._images:
    col = image.anchor._from.col
    row = image.anchor._from.row
    cell_name = f"{get_column_letter(col)}{row + 1}"
    
    filename = None
    for c in range(1, 40):
        val = str(ws.cell(row=row+1, column=c).value)
        if val and (val.lower().endswith('.jpg') or val.lower().endswith('.jpeg') or val.lower().endswith('.png')):
            filename = val
            break
            
    if not filename:
        filename = f"cand_img_{cell_name}.png"
        
    img_data = image._data()
    img = Image.open(io.BytesIO(img_data))
    
    if img.mode in ("RGBA", "P") and (filename.lower().endswith('.jpg') or filename.lower().endswith('.jpeg')):
        img = img.convert("RGB")
        
    save_path = os.path.join('assets/images/candidates', filename)
    img.save(save_path)
    count += 1

print(f"Extracted {count} images from Candidates sheet.")
